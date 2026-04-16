"""XLSX extractor — renders spreadsheets as tab-separated text per sheet."""
from __future__ import annotations

from io import BytesIO
from typing import List, Set

from app.services.extractors.base import BaseExtractor, ExtractResult


class XlsxExtractor(BaseExtractor):
    """Extracts text from XLSX files, rendering each sheet as tab-separated rows."""

    @property
    def extensions(self) -> Set[str]:
        return {"xlsx"}

    @property
    def kind(self) -> str:
        return "xlsx"

    def extract(self, data: bytes, filename: str) -> ExtractResult:
        warnings: List[str] = []
        text = ""
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
            out_lines: List[str] = []
            for ws in wb.worksheets:
                out_lines.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    out_lines.append(
                        "\t".join("" if v is None else str(v) for v in row)
                    )
                out_lines.append("")
            text = "\n".join(out_lines).strip()
        except Exception as e:
            warnings.append(f"XLSX extraction failed: {e!r}")
            text = ""

        return ExtractResult(text=text, kind="xlsx", meta={}, warnings=warnings)
