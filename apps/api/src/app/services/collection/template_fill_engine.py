"""Fill XLSX/XLSM templates from their deterministic marker contract."""
from __future__ import annotations

from copy import copy
import io
import re
from dataclasses import dataclass, field
from typing import Any

from app.services.collection.template_contract import (
    DocumentFormat,
    Orientation,
    TemplateContract,
    ValidationReport,
)
from app.services.collection.template_layout_parser import _parse_placeholder_expr


@dataclass
class FillResult:
    success: bool
    content: bytes | None = None
    error: str | None = None
    filled_scalars: list[str] = field(default_factory=list)
    filled_tables: list[str] = field(default_factory=list)
    missing_scalars: list[str] = field(default_factory=list)
    missing_tables: list[str] = field(default_factory=list)
    validation: ValidationReport | None = None


class TemplateFillEngine:
    """Copies only explicit technical marker rows; it never infers a table."""

    def __init__(self, contract: TemplateContract):
        self.contract = contract

    def fill(self, template_bytes: bytes, values: dict[str, Any], filename: str, *, assume_valid: bool = False) -> FillResult:
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            return FillResult(False, error="Only .xlsx and .xlsm templates are supported")
        if self.contract.format not in (None, DocumentFormat.EXCEL):
            return FillResult(False, error="Template contract is not an Excel contract")
        report = self.contract.validate_generated_values(values)
        if not assume_valid and not report.ok:
            return FillResult(False, error=f"Validation failed: {report.errors}", validation=report)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=filename.lower().endswith(".xlsm"))
        except Exception as exc:
            return FillResult(False, error=f"Failed to load Excel: {exc}")

        normalized = self.contract.normalize_values(values)
        scalar_values = {item.key: _lookup(normalized, item.key) for item in self.contract.scalar_fields()}
        used_scalars: set[str] = set()
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        value, used = _substitute(cell.value, scalar_values)
                        cell.value = value
                        used_scalars.update(used)

        filled_tables: list[str] = []
        missing_tables: list[str] = []
        for table in self.contract.table_fields():
            rows = _lookup(normalized, table.key)
            if rows in (None, ""):
                rows = []
            if not isinstance(rows, list):
                return FillResult(False, error=f"Table '{table.key}' must be an array")
            if table.orientation != Orientation.VERTICAL or not table.anchor or not table.anchor.marker:
                return FillResult(False, error=f"Table '{table.key}' has no supported marker-row anchor")
            ok, error = self._fill_marker_rows(wb, table.key, table.anchor.sheet, table.anchor.marker.row, rows)
            if not ok:
                missing_tables.append(table.key)
                return FillResult(False, error=error, missing_tables=missing_tables)
            filled_tables.append(table.key)

        out = io.BytesIO()
        wb.save(out)
        return FillResult(
            True,
            content=out.getvalue(),
            filled_scalars=sorted(used_scalars),
            filled_tables=filled_tables,
            missing_scalars=[x.key for x in self.contract.scalar_fields() if x.key not in used_scalars],
            missing_tables=missing_tables,
            validation=report,
        )

    def _fill_marker_rows(self, wb: Any, key: str, sheet_name: str | None, marker_row: int, rows: list[dict[str, Any]]) -> tuple[bool, str | None]:
        if not sheet_name or sheet_name not in wb.sheetnames:
            return False, f"Table '{key}' marker sheet was not found"
        ws = wb[sheet_name]
        if marker_row > ws.max_row:
            return False, f"Table '{key}' marker row was not found"
        source = list(ws.iter_rows(min_row=marker_row, max_row=marker_row))[0]
        values = [cell.value for cell in source]
        if not any(isinstance(value, str) and f"{{{{{key}" in value for value in values):
            return False, f"Table '{key}' marker no longer matches the template"
        if not rows:
            ws.delete_rows(marker_row, 1)
            return True, None
        height = ws.row_dimensions[marker_row].height
        hidden = ws.row_dimensions[marker_row].hidden
        for index, row in enumerate(rows):
            target = marker_row + index
            if index:
                ws.insert_rows(target)
                self._copy_row(ws, source, target, height, hidden)
            self._write_row(ws, source, values, target, key, row)
        return True, None

    @staticmethod
    def _copy_row(ws: Any, source: list[Any], target_row: int, height: Any, hidden: Any) -> None:
        ws.row_dimensions[target_row].height = height
        ws.row_dimensions[target_row].hidden = hidden
        for cell in source:
            target = ws.cell(target_row, cell.column)
            if cell.has_style:
                target._style = copy(cell._style)
            target.number_format = copy(cell.number_format)
            target.font = copy(cell.font)
            target.fill = copy(cell.fill)
            target.border = copy(cell.border)
            target.alignment = copy(cell.alignment)
            target.protection = copy(cell.protection)

    @staticmethod
    def _write_row(ws: Any, source: list[Any], source_values: list[Any], target_row: int, table_key: str, row: dict[str, Any]) -> None:
        values = {f"{table_key}.{path}": value for path, value in _flatten(row).items()}
        for cell, raw in zip(source, source_values):
            target = ws.cell(target_row, cell.column)
            target.value = _substitute(raw, values)[0] if isinstance(raw, str) else raw


def _substitute(text: str, values: dict[str, Any]) -> tuple[str, set[str]]:
    used: set[str] = set()
    def replace(match: re.Match[str]) -> str:
        parsed = _parse_placeholder_expr(match.group(1))
        if not parsed:
            return match.group(0)
        key = parsed[0]
        if key not in values:
            return match.group(0)
        value = _lookup(values, key)
        if value not in (None, ""):
            used.add(key)
            return str(value)
        # A token is technical data, not user-visible fallback text.
        return ""
    return re.sub(r"\{\{([^{}]+)\}\}", replace, text), used


def _flatten(value: Any, prefix: str = "") -> dict[str, Any]:
    if not isinstance(value, dict):
        return {prefix: value} if prefix else {}
    result: dict[str, Any] = {}
    for key, nested in value.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        result.update(_flatten(nested, path))
    return result


def _lookup(value: Any, path: str) -> Any:
    if isinstance(value, dict) and path in value:
        return value[path]
    current = value
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current
