"""
File Analyze Tool — inspect the structure of a spreadsheet (Excel or CSV).

The agent uses this BEFORE filling a template to learn:
- Sheet names (for Excel)
- Column headers
- Data types per column
- Sample rows
- Row counts

This helps the agent understand what fields the template expects.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, ClassVar, Dict, List, Optional

from app.agents.context import ToolContext, ToolResult
from app.agents.handlers.versioned_tool import VersionedTool, register_tool, tool_version
from app.core.logging import get_logger

logger = get_logger(__name__)

_MAX_READ_BYTES = 2 * 1024 * 1024
_SAMPLE_ROWS = 5
_MAX_COLS = 100

_FILE_ID_RE = re.compile(
    r"^(?:chatatt_([0-9a-fA-F-]{36})|ragdoc_([0-9a-fA-F-]{36})_(original|canonical)|colexp_([0-9a-fA-F-]{36}))$"
)

_INPUT_SCHEMA_V1 = {
    "type": "object",
    "properties": {
        "file_id": {
            "type": "string",
            "description": (
                "File ID to analyze. Supported: 'chatatt_<uuid>', "
                "'ragdoc_<uuid>_original', 'colexp_<uuid>'. "
                "Use collection.get_document to obtain file_id for a collection document."
            ),
        },
    },
    "required": ["file_id"],
}

_OUTPUT_SCHEMA_V1 = {
    "type": "object",
    "properties": {
        "format": {"type": "string", "description": "'excel' or 'csv'"},
        "file_name": {"type": "string"},
        "sheets": {
            "type": "array",
            "items": {"type": "string"},
            "description": "Sheet names (for Excel). For CSV, contains a single empty string.",
        },
        "columns": {
            "type": "object",
            "description": "Dict mapping sheet name to list of column headers",
        },
        "data_types": {
            "type": "object",
            "description": "Dict mapping 'sheet:col' or 'col' to inferred type: string, number, date, boolean, mixed",
        },
        "sample_rows": {
            "type": "array",
            "items": {"type": "object"},
            "description": "First few rows as dicts keyed by column name",
        },
        "row_counts": {
            "type": "object",
            "description": "Dict mapping sheet name to total row count (approximate for Excel, exact for CSV)",
        },
        "encoding": {"type": "string", "description": "Detected encoding for CSV files"},
    },
}


def _guess_type(values: List[Any]) -> str:
    """Infer the dominant data type of a list of cell values."""
    if not values:
        return "string"

    types = {"int": 0, "float": 0, "date": 0, "bool": 0, "string": 0}
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            types["bool"] += 1
        elif isinstance(v, int):
            types["int"] += 1
        elif isinstance(v, float):
            types["float"] += 1
        else:
            sv = str(v).strip()
            if sv.lower() in {"true", "false", "yes", "no", "1", "0"}:
                types["bool"] += 1
            else:
                try:
                    float(sv)
                    types["float"] += 1
                except ValueError:
                    types["string"] += 1

    max_type = max(types, key=types.get)  # type: ignore[arg-type]
    if max_type in {"int", "float"}:
        return "number"
    return max_type if types[max_type] > 0 else "string"


def _analyze_excel(data: bytes, file_name: str) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is not installed; cannot analyze Excel files")

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets = wb.sheetnames
    columns: Dict[str, List[str]] = {}
    data_types: Dict[str, str] = {}
    sample_rows: List[Dict[str, Any]] = []
    row_counts: Dict[str, int] = {}

    for sheet_name in sheets:
        ws = wb[sheet_name]
        headers: List[str] = []
        rows_iter = iter(ws.rows)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            columns[sheet_name] = []
            row_counts[sheet_name] = 0
            continue

        # Build headers
        headers = []
        for idx, cell in enumerate(first_row):
            if idx >= _MAX_COLS:
                break
            val = cell.value
            if val is None:
                val = f"Column{idx + 1}"
            headers.append(str(val).strip())
        columns[sheet_name] = headers

        rows_data: List[List[Any]] = []
        for row in rows_iter:
            row_vals = [cell.value for cell in row[: len(headers)]]
            rows_data.append(row_vals)

        row_counts[sheet_name] = len(rows_data)

        # Sample rows (from first sheet only, to keep output compact)
        if sheet_name == sheets[0]:
            for rvals in rows_data[:_SAMPLE_ROWS]:
                sample_rows.append({h: v for h, v in zip(headers, rvals)})

        # Data types per column
        for idx, h in enumerate(headers):
            col_vals = [r[idx] if idx < len(r) else None for r in rows_data]
            data_types[f"{sheet_name}:{h}"] = _guess_type(col_vals)

    wb.close()
    return {
        "format": "excel",
        "file_name": file_name,
        "sheets": sheets,
        "columns": columns,
        "data_types": data_types,
        "sample_rows": sample_rows,
        "row_counts": row_counts,
        "encoding": None,
    }


def _analyze_csv(data: bytes, file_name: str) -> Dict[str, Any]:
    # Try UTF-8 first, then common fallbacks
    encodings = ["utf-8", "utf-8-sig", "cp1251", "cp1252", "iso-8859-1"]
    text = None
    used_encoding = "utf-8"
    for enc in encodings:
        try:
            text = data.decode(enc)
            used_encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RuntimeError("Could not decode CSV file with any known encoding")

    reader = csv.reader(io.StringIO(text))
    try:
        headers = [str(h).strip() for h in next(reader)]
    except StopIteration:
        headers = []

    rows_data: List[List[Any]] = []
    for row in reader:
        rows_data.append(row)
        if len(rows_data) >= _SAMPLE_ROWS + 1000:  # safety cap
            break

    sample_rows: List[Dict[str, Any]] = []
    for rvals in rows_data[:_SAMPLE_ROWS]:
        sample_rows.append({h: v for h, v in zip(headers, rvals)})

    data_types: Dict[str, str] = {}
    for idx, h in enumerate(headers):
        col_vals = [r[idx] if idx < len(r) else None for r in rows_data]
        data_types[h] = _guess_type(col_vals)

    return {
        "format": "csv",
        "file_name": file_name,
        "sheets": [""],
        "columns": {"": headers},
        "data_types": data_types,
        "sample_rows": sample_rows,
        "row_counts": {"": len(rows_data)},
        "encoding": used_encoding,
    }


@register_tool
class FileAnalyzeTool(VersionedTool):
    """
    Inspect the structure of a spreadsheet (Excel or CSV).

    Use this BEFORE filling a template to learn what sheets, columns, and data
    types the file contains. Returns sample rows so the agent knows the expected
    format.
    """

    tool_slug: ClassVar[str] = "file.analyze"
    domains: ClassVar[list] = ["system"]
    name: ClassVar[str] = "Analyze Spreadsheet"
    description: ClassVar[str] = (
        "Inspect spreadsheet structure (Excel/CSV): sheets, column headers, data types, sample rows, row counts. "
        "Use BEFORE filling a template to learn what fields it expects. "
        "Accepts any file_id supported by file.read."
    )

    @tool_version(
        version="1.0.0",
        input_schema=_INPUT_SCHEMA_V1,
        output_schema=_OUTPUT_SCHEMA_V1,
        description="Analyze Excel or CSV structure",
    )
    async def v1_0_0(self, ctx: ToolContext, args: Dict[str, Any]) -> ToolResult:
        from app.adapters.s3_client import s3_manager
        from app.core.db import get_session_factory
        from app.repositories.factory import AsyncRepositoryFactory
        from app.services.file_delivery_service import FileDeliveryService

        log = ctx.tool_logger("file.analyze")

        file_id = str(args.get("file_id") or "").strip()
        if not file_id:
            log.error("Missing file_id")
            return ToolResult.fail("Missing 'file_id' argument", logs=log.entries_dict())

        if not _FILE_ID_RE.match(file_id):
            log.error("Invalid file_id format", file_id=file_id)
            return ToolResult.fail(
                f"Invalid file_id format '{file_id}'. "
                f"Expected 'chatatt_<uuid>', 'ragdoc_<uuid>_original', or 'colexp_<uuid>'.",
                logs=log.entries_dict(),
            )

        user_id = ctx.user_id
        if not user_id:
            log.error("Missing user_id in tool context")
            return ToolResult.fail(
                "File analyze requires a user context.",
                logs=log.entries_dict(),
            )

        log.info("Analyzing file", file_id=file_id)

        try:
            session_factory = get_session_factory()
            async with session_factory() as session:
                repo_factory = AsyncRepositoryFactory(
                    session, tenant_id=ctx.tenant_id, user_id=user_id
                )
                service = FileDeliveryService(session, repo_factory)
                resolved = await service.resolve(file_id, owner_id=str(user_id))

                payload = await s3_manager.get_object(resolved.bucket, resolved.key)
                if payload is None:
                    log.error("Failed to load file from storage", file_id=file_id)
                    return ToolResult.fail(
                        f"File '{file_id}' exists in database but could not be loaded from storage.",
                        logs=log.entries_dict(),
                    )

                size_bytes = len(payload)
                if size_bytes > _MAX_READ_BYTES:
                    log.error(
                        "File too large",
                        size_bytes=size_bytes,
                        max_bytes=_MAX_READ_BYTES,
                    )
                    return ToolResult.fail(
                        f"File size ({size_bytes} bytes) exceeds analyze limit of {_MAX_READ_BYTES} bytes.",
                        logs=log.entries_dict(),
                    )

                # Detect format by extension
                fname = resolved.file_name or ""
                ext = ""
                if "." in fname:
                    ext = fname.rsplit(".", 1)[-1].strip().lower()

                if ext in {"xlsx", "xls", "xlsm"}:
                    result = _analyze_excel(payload, fname)
                elif ext in {"csv", "tsv"}:
                    result = _analyze_csv(payload, fname)
                else:
                    log.error("Unsupported file format for analysis", ext=ext)
                    return ToolResult.fail(
                        f"File format '{ext or 'unknown'}' is not supported for analysis. "
                        f"Supported: xlsx, xls, xlsm, csv, tsv.",
                        logs=log.entries_dict(),
                    )

                log.info(
                    "File analyzed",
                    file_id=file_id,
                    format=result["format"],
                    sheets=len(result["sheets"]),
                )
                return ToolResult.ok(
                    data=result,
                    message=f"Analyzed '{fname}' ({result['format']}, {len(result['sheets'])} sheet(s)).",
                    logs=log.entries_dict(),
                )
        except Exception as exc:
            logger.error("File analyze failed: %s", exc, exc_info=True)
            log.error("File analyze failed", error=str(exc))
            return ToolResult.fail(
                f"Failed to analyze file: {exc}",
                logs=log.entries_dict(),
            )
